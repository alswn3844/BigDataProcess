#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
sheet_ranges = wb['Sheet1']

column = ['A', 'B', 'C', 'D', 'E', 'F']
midterm = []
final = []
homework = []
attendance = []
total = []
grades = ['A+', 'A0', 'B+', 'B0', 'C+', 'C0', 'F']
numofPeople = len(sheet_ranges['C'])-1
for i in column: 
        for j in range(2, numofPeople+2):
                if i == 'C': 
                        midterm.append(sheet_ranges[i+str(j)].value)
                elif i == 'D': 
                        final.append(sheet_ranges[i+str(j)].value)
                elif i == 'E': 
                        homework.append(sheet_ranges[i+str(j)].value)
                elif i == 'F': 
                        attendance.append(sheet_ranges[i+str(j)].value)

for i in range(0, numofPeople):
        total.append(float(midterm[i] * 0.3 + final[i] * 0.35 + homework[i] * 0.34 + attendance[i]))

for i in range(0, numofPeople):
        sheet_ranges.cell(column = 7, row = i+2, value = total[i])

A = int(numofPeople*0.3)
B = int(numofPeople*0.7)-A
C = numofPeople-A-B
biggest = 1

for i in range(1, numofPeople+1):
	for j in range(2, numofPeople+2):
		if sheet_ranges['H'+str(j)].value != None:
			continue
		if biggest == 1 or sheet_ranges['G'+str(biggest)].value < sheet_ranges['G'+str(j)].value:
			biggest = j 
	if i<=A/2:
		sheet_ranges['H'+str(biggest)] = 'A+'
	elif A/2 < i and i <= A:
		sheet_ranges['H'+str(biggest)] = 'A0'
	elif A < i and i <= A+(B/2):
		sheet_ranges['H'+str(biggest)] = 'B+'
	elif A+(B/2) < i and i <= A+B:
		sheet_ranges['H'+str(biggest)] = 'B0'
	elif A+B < i and i <= A+B+(C/2):
		sheet_ranges['H'+str(biggest)] = 'C+'
	else:
		sheet_ranges['H'+str(biggest)] = 'C0'
	biggest = 1
		
wb.save('student.xlsx')
